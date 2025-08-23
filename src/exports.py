import csv
import io
import json
from datetime import datetime
from sqlalchemy.orm import Session
from .database import ScrapedJob
import openpyxl
from openpyxl.styles import Font, PatternFill

def export_to_csv(jobs: list) -> str:
    """Exporta vagas para CSV"""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Título', 'Link', 'Fonte', 'Data Coleta'])
    
    for job in jobs:
        writer.writerow([
            job.title,
            job.link,
            job.source,
            job.scraped_at.strftime('%d/%m/%Y %H:%M')
        ])
    
    return output.getvalue()

def export_to_excel(jobs: list) -> bytes:
    """Exporta vagas para Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vagas Coletadas"
    
    # Cabeçalhos
    headers = ['Título', 'Link', 'Fonte', 'Data Coleta']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Dados
    for row, job in enumerate(jobs, 2):
        ws.cell(row=row, column=1, value=job.title)
        ws.cell(row=row, column=2, value=job.link)
        ws.cell(row=row, column=3, value=job.source)
        ws.cell(row=row, column=4, value=job.scraped_at.strftime('%d/%m/%Y %H:%M'))
    
    # Ajustar largura das colunas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Salvar em bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

def export_to_json(jobs: list) -> str:
    """Exporta vagas para JSON"""
    data = []
    for job in jobs:
        data.append({
            'title': job.title,
            'link': job.link,
            'source': job.source,
            'scraped_at': job.scraped_at.isoformat()
        })
    
    return json.dumps(data, indent=2, ensure_ascii=False)